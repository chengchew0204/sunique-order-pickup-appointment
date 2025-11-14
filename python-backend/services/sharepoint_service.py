import msal
import requests
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

class SharePointService:
    def __init__(self, config):
        self.config = config
        self.client_id = config.get('CLIENT_ID')
        self.client_secret = config.get('CLIENT_SECRET')
        self.tenant_id = config.get('TENANT_ID')
        self.site_id = config.get('SHAREPOINT_SITE_ID')
        self.site_url = config.get('SHAREPOINT_SITE_URL')
        
        self.authority = f"https://login.microsoftonline.com/{self.tenant_id}"
        self.scope = ['https://graph.microsoft.com/.default']
        
        self.app = msal.ConfidentialClientApplication(
            self.client_id,
            authority=self.authority,
            client_credential=self.client_secret
        )
    
    def get_access_token(self):
        """Get Microsoft Graph API access token"""
        result = self.app.acquire_token_silent(self.scope, account=None)
        
        if not result:
            result = self.app.acquire_token_for_client(scopes=self.scope)
        
        if "access_token" in result:
            return result['access_token']
        else:
            raise Exception(f"Failed to acquire token: {result.get('error_description', 'Unknown error')}")
    
    def get_file_content(self, file_path):
        """Download file content from SharePoint"""
        try:
            token = self.get_access_token()
            headers = {'Authorization': f'Bearer {token}'}
            
            # Always look up site ID from URL to ensure correct format
            # Microsoft Graph site IDs need to be in format: hostname,siteId,webId
            site_id = self._get_site_id()
            
            url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:{file_path}:/content"
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                return response.content
            else:
                raise Exception(f"Failed to fetch file: {response.status_code} - {response.text}")
                
        except Exception as e:
            print(f"Error fetching file from SharePoint: {e}")
            raise
    
    def upload_file_content(self, file_path, content):
        """Upload file content to SharePoint with retry logic"""
        max_retries = 5
        
        for attempt in range(1, max_retries + 1):
            try:
                token = self.get_access_token()
                headers = {'Authorization': f'Bearer {token}'}
                
                # Always look up site ID from URL to ensure correct format
                site_id = self._get_site_id()
                
                url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:{file_path}:/content"
                response = requests.put(url, headers=headers, data=content)
                
                if response.status_code in [200, 201]:
                    if attempt > 1:
                        print(f"Upload succeeded on attempt {attempt}")
                    return True
                elif response.status_code == 423:  # Locked
                    if attempt < max_retries:
                        wait_time = 2 ** (attempt - 1)
                        print(f"File locked, retrying in {wait_time}s... (attempt {attempt}/{max_retries})")
                        import time
                        time.sleep(wait_time)
                        continue
                else:
                    raise Exception(f"Failed to upload file: {response.status_code} - {response.text}")
                    
            except Exception as e:
                if attempt == max_retries:
                    print(f"Error uploading file to SharePoint: {e}")
                    raise
                else:
                    wait_time = 2 ** (attempt - 1)
                    print(f"Upload failed, retrying in {wait_time}s... (attempt {attempt}/{max_retries})")
                    import time
                    time.sleep(wait_time)
        
        raise Exception(f"Failed to upload file after {max_retries} attempts")
    
    def _get_site_id(self):
        """Look up SharePoint site ID from URL"""
        try:
            token = self.get_access_token()
            headers = {'Authorization': f'Bearer {token}'}
            
            # Extract hostname and site path from URL
            import re
            match = re.match(r'https?://([^/]+)/sites/([^/]+)', self.site_url)
            if not match:
                raise Exception('Invalid SharePoint site URL format')
            
            hostname = match.group(1)
            site_path = match.group(2)
            
            url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:/sites/{site_path}"
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                return response.json()['id']
            else:
                raise Exception(f"Failed to get site ID: {response.status_code} - {response.text}")
                
        except Exception as e:
            print(f"Error getting site ID: {e}")
            raise
    
    def parse_excel_file(self, content):
        """Parse Excel file content and return as list of dictionaries"""
        try:
            # Load Excel file from bytes
            excel_file = BytesIO(content)
            workbook = load_workbook(excel_file, data_only=True)
            sheet = workbook.active
            
            # Find header row containing "Ready Order Number"
            header_row_index = None
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                if any(cell and str(cell).strip() == 'Ready Order Number' for cell in row):
                    header_row_index = i
                    break
            
            if not header_row_index:
                # Default to first row if not found
                header_row_index = 1
            
            # Read data using pandas with correct header row
            excel_file.seek(0)
            df = pd.read_excel(excel_file, header=header_row_index - 1)
            
            # Convert to list of dictionaries
            records = df.to_dict('records')
            
            # Clean up NaN values
            for record in records:
                for key, value in record.items():
                    if pd.isna(value):
                        record[key] = ''
            
            return records
            
        except Exception as e:
            print(f"Error parsing Excel file: {e}")
            raise
    
    def parse_csv_file(self, content):
        """Parse CSV file content and return as list of dictionaries"""
        try:
            csv_file = BytesIO(content)
            df = pd.read_csv(csv_file)
            
            # Convert to list of dictionaries
            records = df.to_dict('records')
            
            # Clean up NaN values
            for record in records:
                for key, value in record.items():
                    if pd.isna(value):
                        record[key] = ''
            
            return records
            
        except Exception as e:
            print(f"Error parsing CSV file: {e}")
            raise
    
    def records_to_csv_bytes(self, records, fieldnames):
        """Convert list of dictionaries to CSV bytes"""
        import csv
        
        output = BytesIO()
        output.write(b'\xef\xbb\xbf')  # UTF-8 BOM
        
        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            lineterminator='\n',
            extrasaction='ignore'
        )
        
        writer.writeheader()
        writer.writerows(records)
        
        return output.getvalue()

